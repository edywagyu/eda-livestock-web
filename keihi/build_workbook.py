"""Build a single Excel workbook for the accountant submission."""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "個人立替経費_税理士提出.xlsx"

SB_CSV = HERE / "starbucks" / "starbucks_receipts.csv"
AMEX_CONF = HERE / "amex" / "amex_business_confident.csv"
AMEX_UNC = HERE / "amex" / "amex_uncertain.csv"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
SECTION_FONT = Font(bold=True, size=12)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def add_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("0_サマリ・依頼内容", 0)
    ws["A1"] = "個人立替経費 ご精算依頼"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "江田畜産株式会社　代表取締役　江田友輝"
    ws["A3"] = "作成日：2026年5月25日"

    ws["A5"] = "■ 概要"
    ws["A5"].font = SECTION_FONT
    ws["A5"].fill = SECTION_FILL
    ws["A6"] = "代表者が個人カード（AmEx）・個人メールで立替えていた業務関連支出の精算依頼。"
    ws["A7"] = "第3期に取りこぼしたスタバ経費も含めて取りまとめ。"

    ws["A9"] = "■ 集計結果"
    ws["A9"].font = SECTION_FONT
    ws["A9"].fill = SECTION_FILL

    headers = ["区分", "対象期間", "勘定科目", "件数", "金額（税込）"]
    ws.append([])
    for c, h in enumerate(headers, 1):
        ws.cell(row=11, column=c, value=h)
    style_header(ws, 11, len(headers))

    data = [
        ["第3期分", "2024/4〜2025/3", "会議費（スタバ）", 30, 18638],
        ["第4期分", "2025/4〜2026/3", "旅費交通費", 64, 144644],
        ["第4期分", "2025/4〜2026/3", "通信費", 52, 136859],
        ["第4期分", "2025/4〜2026/3", "会議費", 51, 64657],
    ]
    for row in data:
        ws.append(row)

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(r[3] for r in data)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(r[4] for r in data)).font = Font(bold=True)
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="FFF2CC")

    for r in range(12, total_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            if c >= 4:
                ws.cell(row=r, column=c).number_format = "#,##0"

    ws[f"A{total_row + 2}"] = "■ 仕訳案（ご判断お願いします）"
    ws[f"A{total_row + 2}"].font = SECTION_FONT
    ws[f"A{total_row + 2}"].fill = SECTION_FILL

    explain = [
        "",
        "案A：役員借入金として処理（推奨）",
        "  （借）旅費交通費  144,644 ／（貸）役員借入金  346,160",
        "  （借）通信費      136,859",
        "  （借）会議費       64,657",
        "  → 後日、会社→代表者へ振込で精算",
        "",
        "案B：未払金として処理",
        "  （借）旅費交通費  144,644 ／（貸）未払金  346,160",
        "  （借）通信費      136,859",
        "  （借）会議費       64,657",
        "  → 後日精算時に未払金消込",
    ]
    for line in explain:
        ws.append([line])

    last = ws.max_row + 2
    ws.cell(row=last, column=1, value="■ ご確認いただきたい事項（5点）").font = SECTION_FONT
    ws.cell(row=last, column=1).fill = SECTION_FILL
    questions = [
        "1. 第3期分（¥18,638）の追加計上方法（修正申告 or 第4期での前期損益修正）",
        "2. 個人立替の精算処理 → 案A/案B どちらが適切か",
        "3. インボイス少額特例の当社適用可否（基準期間課税売上1億円以下が条件）",
        "4. 保留分597件（コンビニ等）は一括除外で良いか、按分計上すべきものがあるか",
        "5. 証憑保管：AmEx明細CSV＋メール領収書のみで7年保存要件を満たすか",
    ]
    for q in questions:
        ws.append([q])

    last2 = ws.max_row + 2
    ws.cell(row=last2, column=1, value="■ 証憑（エビデンス）の状況").font = SECTION_FONT
    ws.cell(row=last2, column=1).fill = SECTION_FILL
    ev = [
        "【揃っているもの】",
        "  ・スタバ Mobile Order 領収書メール（インボイス番号入り・完全証憑）",
        "  ・AmEx 利用明細CSV（過去13ヶ月分）",
        "  ・AmEx 月次請求書PDF（要望あれば追加取得可）",
        "",
        "【揃っていないもの／代用案】",
        "  ・コンビニ・カフェの紙レシート → AmEx明細で代用希望",
        "  ・海外出張の現地レシート → クレカ明細＋メール旅程で代用希望",
    ]
    for e in ev:
        ws.append([e])

    last3 = ws.max_row + 2
    ws.cell(row=last3, column=1, value="■ シート一覧").font = SECTION_FONT
    ws.cell(row=last3, column=1).fill = SECTION_FILL
    sheets = [
        "  0_サマリ・依頼内容（このシート）",
        "  1_スタバ3期_全明細（30件）",
        "  2_スタバ3期_月別",
        "  3_AmEx4期_業務確実分（168件）",
        "  4_AmEx4期_月別科目別",
        "  5_AmEx4期_保留判断分（597件・要相談）",
    ]
    for s in sheets:
        ws.append([s])

    autosize(ws, {1: 60, 2: 20, 3: 25, 4: 10, 5: 15})


def add_sb_detail_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("1_スタバ3期_全明細")
    rows = read_csv(SB_CSV)
    headers = list(rows[0].keys())
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for r in rows:
        ws.append([r[h] for h in headers])

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    total = sum(int(r["総合計"]) for r in rows)
    count_col = headers.index("総合計") + 1
    ws.cell(row=total_row, column=count_col, value=total).font = Font(bold=True)
    ws.cell(row=total_row, column=count_col).fill = PatternFill("solid", fgColor="FFF2CC")

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            if headers[c - 1] in ("本体合計", "消費税", "総合計"):
                ws.cell(row=r, column=c).number_format = "#,##0"

    widths = {1: 18, 2: 38, 3: 60, 4: 12, 5: 8, 6: 10, 7: 12, 8: 18, 9: 18, 10: 35, 11: 20}
    autosize(ws, widths)
    ws.freeze_panes = "A2"


def add_sb_monthly_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("2_スタバ3期_月別")
    rows = read_csv(SB_CSV)
    buckets: dict[str, dict] = defaultdict(lambda: {"件数": 0, "本体": 0, "税": 0, "総額": 0})
    for r in rows:
        ym = r["注文日時"][:7]
        b = buckets[ym]
        b["件数"] += 1
        b["本体"] += int(r["本体合計"])
        b["税"] += int(r["消費税"])
        b["総額"] += int(r["総合計"])

    headers = ["年月", "件数", "本体合計", "消費税", "総合計"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for ym in sorted(buckets):
        b = buckets[ym]
        ws.append([ym, b["件数"], b["本体"], b["税"], b["総額"]])
    total = {k: sum(v[k] for v in buckets.values()) for k in ["件数", "本体", "税", "総額"]}
    ws.append(["合計", total["件数"], total["本体"], total["税"], total["総額"]])

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            if c >= 3:
                ws.cell(row=r, column=c).number_format = "#,##0"
        if r == ws.max_row:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
                ws.cell(row=r, column=c).font = Font(bold=True)

    autosize(ws, {1: 12, 2: 10, 3: 14, 4: 12, 5: 14})


def add_amex_detail_sheet(wb: Workbook, name: str, path: Path, with_total: bool = True) -> None:
    ws = wb.create_sheet(name)
    rows = read_csv(path)
    if not rows:
        ws.append(["データなし"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for r in rows:
        out_row = []
        for h in headers:
            val = r[h]
            if h == "金額":
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            out_row.append(val)
        ws.append(out_row)

    if with_total and "金額" in headers:
        amount_col = headers.index("金額") + 1
        total_row = ws.max_row + 1
        total = sum(int(r["金額"]) for r in rows if r["金額"] and r["金額"].lstrip("-").isdigit())
        ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
        ws.cell(row=total_row, column=amount_col, value=total).font = Font(bold=True)
        ws.cell(row=total_row, column=amount_col).fill = PatternFill("solid", fgColor="FFF2CC")

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            if headers[c - 1] == "金額":
                ws.cell(row=r, column=c).number_format = "#,##0"

    widths = {1: 13, 2: 42, 3: 12, 4: 18, 5: 18, 6: 35}
    autosize(ws, widths)
    ws.freeze_panes = "A2"


def add_amex_monthly_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("4_AmEx4期_月別科目別")
    rows = read_csv(AMEX_CONF)
    buckets: dict[tuple[str, str], dict] = defaultdict(lambda: {"件数": 0, "金額": 0})
    for r in rows:
        ym = r["ご利用日"].replace("/", "-")[:7]
        amount = int(r["金額"]) if r["金額"].lstrip("-").isdigit() else 0
        key = (ym, r["勘定科目候補"])
        buckets[key]["件数"] += 1
        buckets[key]["金額"] += amount

    headers = ["年月", "勘定科目", "件数", "金額"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for key in sorted(buckets):
        ym, kamoku = key
        b = buckets[key]
        ws.append([ym, kamoku, b["件数"], b["金額"]])

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=sum(b["件数"] for b in buckets.values())).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(b["金額"] for b in buckets.values())).font = Font(bold=True)
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="FFF2CC")

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            if c >= 3:
                ws.cell(row=r, column=c).number_format = "#,##0"

    autosize(ws, {1: 12, 2: 18, 3: 10, 4: 14})


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    add_summary_sheet(wb)
    add_sb_detail_sheet(wb)
    add_sb_monthly_sheet(wb)
    add_amex_detail_sheet(wb, "3_AmEx4期_業務確実分", AMEX_CONF)
    add_amex_monthly_sheet(wb)
    add_amex_detail_sheet(wb, "5_AmEx4期_保留判断分", AMEX_UNC)

    wb.save(OUT)
    print(f"出力: {OUT}")


if __name__ == "__main__":
    main()
